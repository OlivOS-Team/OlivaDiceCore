# -*- encoding: utf-8 -*-
'''
_______________________    _________________________________________
__  __ \__  /____  _/_ |  / /__    |__  __ \___  _/_  ____/__  ____/
_  / / /_  /  __  / __ | / /__  /| |_  / / /__  / _  /    __  __/   
/ /_/ /_  /____/ /  __ |/ / _  ___ |  /_/ /__/ /  / /___  _  /___   
\____/ /_____/___/  _____/  /_/  |_/_____/ /___/  \____/  /_____/   

@File      :   drawCard.py
@Author    :   lunzhiPenxil仑质
@Contact   :   lunzhipenxil@gmail.com
@License   :   AGPL
@Copyright :   (C) 2020-2021, OlivOS-Team
@Desc      :   None
'''

import OlivOS
import OlivaDiceCore

import os
import json
#import yaml
#import openpyxl
import re
import traceback
import random
import codecs
import traceback

# 兼容OlivOS 0.10.2及以下版本
try:
    import openpyxl
except:
    pass
try:
    import yaml
except:
    pass
try:
    import pyjson5
except:
    pass

dictReMappingDrawFormat = {
    'player': 'tName'
}

def reMappingDrawFormat(data:str):
    res = data
    for key in dictReMappingDrawFormat:
        res = res.replace('{%s}' % key, '{%s}' % dictReMappingDrawFormat[key])
    return res

def formatUTF8WithBOM(data:bytes):
    res = data
    if res[:3] == codecs.BOM_UTF8:
        res = res[3:]
    return res

def initDeckHelp(bot_info_dict):
    deck_name_tmp_list = list(OlivaDiceCore.drawCardData.dictDeckTemp.keys())
    deck_name_tmp_list = [
        deck_name_tmp_list_this
        for deck_name_tmp_list_this in deck_name_tmp_list
        if type(deck_name_tmp_list_this) == str and not deck_name_tmp_list_this.startswith('_')
    ]
    deck_name_tmp_list_str = '/'.join(deck_name_tmp_list)
    for bot_hash in bot_info_dict:
        # 应用重定向逻辑（读取牌堆数据时使用重定向后的botHash）
        redirected_bot_hash = OlivaDiceCore.userConfig.getRedirectedBotHash(bot_hash)
        
        helpdoc_patch = {}
        flag_drawListMode = OlivaDiceCore.console.getConsoleSwitchByHash(
            'drawListMode',
            bot_hash
        )
        if flag_drawListMode == 0:
            pass
        elif flag_drawListMode == 1:
            if redirected_bot_hash in OlivaDiceCore.drawCardData.dictDeck and type(OlivaDiceCore.drawCardData.dictDeck[redirected_bot_hash]) == dict:
                deck_name_list = list(OlivaDiceCore.drawCardData.dictDeck[redirected_bot_hash].keys())
                deck_name_list = [
                    deck_name_list_this
                    for deck_name_list_this in deck_name_list
                    if type(deck_name_list_this) == str and not deck_name_list_this.startswith('_')
                ]
                deck_name_extend_list = [
                    deck_name_list_this
                    for deck_name_list_this in deck_name_list
                    if deck_name_list_this not in deck_name_tmp_list
                ]
                helpdoc_patch = {
                    '全牌堆列表': '/'.join(deck_name_list),
                    '内置牌堆': deck_name_tmp_list_str,
                    '扩展牌堆': '/'.join(deck_name_extend_list)
                }
        elif flag_drawListMode == 2:
            if redirected_bot_hash in OlivaDiceCore.drawCardData.dictDeckIndex and type(OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash]) == dict:
                deck_name_tmp_list_index = {'内置牌堆': '内置牌堆:\n%s' % deck_name_tmp_list_str}
                deck_name_list_index = {
                    '扩展牌堆 %s' % deck_name_list_this: '%s:\n%s' % (
                        deck_name_list_this,
                        '/'.join([
                            deck_name_list_this_this
                            for deck_name_list_this_this in OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash][deck_name_list_this]
                            if type(deck_name_list_this_this) == str and not deck_name_list_this_this.startswith('_')
                        ])
                    )
                    for deck_name_list_this in OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash]
                    if type(deck_name_list_this) == str and type(OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash][deck_name_list_this]) == list
                }
                deck_name_tmp_list_index_check = [
                    deck_name_tmp_list_index_key
                    for deck_name_tmp_list_index_key in deck_name_tmp_list_index
                    if type(deck_name_tmp_list_index_key) == str and type(deck_name_tmp_list_index[deck_name_tmp_list_index_key]) == str
                ]
                deck_name_list_index_check = [
                    deck_name_list_this
                    for deck_name_list_this in OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash]
                    if type(deck_name_list_this) == str and type(OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash][deck_name_list_this]) == list
                ]
                helpdoc_patch.update(deck_name_tmp_list_index)
                helpdoc_patch.update(deck_name_list_index)
                helpdoc_patch.update({
                    '全牌堆列表': '/'.join(deck_name_tmp_list_index_check + deck_name_list_index_check),
                    '扩展牌堆': '/'.join(deck_name_list_index_check)
                })
        elif flag_drawListMode == 3:
            if redirected_bot_hash in OlivaDiceCore.drawCardData.dictDeckIndex and type(OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash]) == dict:
                deck_name_tmp_list_index = ['内置牌堆:\n%s' % deck_name_tmp_list_str]
                deck_name_list_index = [
                    '%s:\n%s' % (
                        deck_name_list_this,
                        '/'.join([
                            deck_name_list_this_this
                            for deck_name_list_this_this in OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash][deck_name_list_this]
                            if type(deck_name_list_this_this) == str and not deck_name_list_this_this.startswith('_')
                        ])
                    )
                    for deck_name_list_this in OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash]
                    if type(OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash][deck_name_list_this]) == list
                ]
                helpdoc_patch = {
                    '全牌堆列表': '\n'.join(deck_name_tmp_list_index + deck_name_list_index),
                    '内置牌堆': '\n'.join(deck_name_tmp_list_index),
                    '扩展牌堆': '\n'.join(deck_name_list_index)
                }
        if bot_hash in OlivaDiceCore.helpDocData.dictHelpDoc:
            OlivaDiceCore.helpDocData.dictHelpDoc[bot_hash].update(helpdoc_patch)

def setDeckIndex(bot_hash:str, deck_name:str, deck_data:dict):
    if type(bot_hash) == str and type(deck_name) == str and type(deck_data) == dict:
        # 应用重定向逻辑
        redirected_bot_hash = OlivaDiceCore.userConfig.getRedirectedBotHash(bot_hash)
        
        if redirected_bot_hash not in OlivaDiceCore.drawCardData.dictDeckIndex:
            OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash] = {}
        if deck_name not in OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash]:
            OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash][deck_name] = []
        deck_data_list = list(deck_data.keys())
        deck_data_list = [
            deck_data_list_this
            for deck_data_list_this in deck_data_list
            if type(deck_data_list_this) == str and not deck_data_list_this.startswith('_')
        ]
        OlivaDiceCore.drawCardData.dictDeckIndex[redirected_bot_hash][deck_name] = deck_data_list

# 清空牌堆数据
def cleanDeck():
    for botInfo in OlivaDiceCore.drawCardData.dictDeck:
        OlivaDiceCore.drawCardData.dictDeck[botInfo] = {}
    for botInfo in OlivaDiceCore.drawCardData.dictDeckIndex:
        OlivaDiceCore.drawCardData.dictDeckIndex[botInfo] = {}

# 重载牌堆数据
# 目前不会清理牌堆帮助文档数据，所以如果发生牌堆减量将会导致牌堆帮助文档数据残留
# 但是我认为这个问题不明显，用户应该不在意 By lunzhiPenxil
def reloadDeck():
    if OlivaDiceCore.data.global_Proc != None \
    and 'bot_info_dict' in OlivaDiceCore.data.global_Proc.Proc_data:
        bot_info_dict = OlivaDiceCore.data.global_Proc.Proc_data['bot_info_dict']
        cleanDeck()
        initDeck(bot_info_dict = bot_info_dict)

# 根除指定牌堆
def removeDeck(deckName:str, botHash:str):
    checkDict = {
        'deckclassic': ['.json', '.json5'],
        'deckyaml': ['', '.yaml'],
        'deckexcel': ['.xlsx', '.xls']
    }
    for deck_type in ['deckclassic', 'deckyaml', 'deckexcel']:
        for dfix in checkDict.get(deck_type, []):
            deck_path = os.path.join('plugin', 'data', 'OlivaDice', botHash, 'extend', deck_type, deckName + dfix)
            if os.path.exists(deck_path):
                try:
                    os.remove(deck_path)
                except Exception as e:
                    traceback.print_exc()

# 初始化牌堆数据
def initDeck(bot_info_dict):
    dictTValue = OlivaDiceCore.msgCustom.dictTValue.copy()
    dictStrConst = OlivaDiceCore.msgCustom.dictStrConst
    dictGValue = OlivaDiceCore.msgCustom.dictGValue
    dictTValue.update(dictGValue)
    obj_Deck_this_count_total_init = 0
    for bot_info_dict_this in bot_info_dict:
        OlivaDiceCore.drawCardData.dictDeck[bot_info_dict_this] = OlivaDiceCore.drawCardData.dictDeckTemp.copy()
        obj_Deck_this_count_total_init = len(OlivaDiceCore.drawCardData.dictDeck[bot_info_dict_this])
        OlivaDiceCore.drawCardData.dictDeckIndex[bot_info_dict_this] = {}
    releaseDir(OlivaDiceCore.data.dataDirRoot + '/unity')
    releaseDir(OlivaDiceCore.data.dataDirRoot + '/unity/extend')
    releaseDir(OlivaDiceCore.data.dataDirRoot + '/unity/extend/deckclassic')
    customDeckDir = OlivaDiceCore.data.dataDirRoot + '/unity/extend/deckclassic'
    fileDeckList = os.listdir(customDeckDir)
    obj_Deck_this = None
    obj_Deck_this_count = 0
    obj_Deck_this_count_total = 0
    dictTValue['tName'] = '全局'
    botHash = None
    # 全局 json 牌堆
    for fileDeckList_this in fileDeckList:
        customDeckFile = fileDeckList_this
        customDeckPath = customDeckDir + '/' + customDeckFile
        obj_Deck_this = None
        try:
            with open(customDeckPath, 'rb') as customDeckPath_f:
                customDeckPath_fs = formatUTF8WithBOM(customDeckPath_f.read()).decode('utf-8')
                try:
                    obj_Deck_this = pyjson5.loads(customDeckPath_fs)
                except:
                    obj_Deck_this = json.loads(customDeckPath_fs)
        except:
            dictTValue['tInitDataName'] = customDeckFile
            OlivaDiceCore.msgReply.globalLog(
                3,
                OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckDataError'], dictTValue),
                [
                    ('OlivaDice', 'default'),
                    ('Init', 'default')
                ]
            )
        if obj_Deck_this != None:
            for bot_info_dict_this in OlivaDiceCore.drawCardData.dictDeck:
                botHash = bot_info_dict_this
                OlivaDiceCore.drawCardData.dictDeck[botHash].update(obj_Deck_this)
                if customDeckFile.endswith('.json5'):
                    customDeckFile = customDeckFile.rstrip('.json5')
                elif customDeckFile.endswith('.json'):
                    customDeckFile = customDeckFile.rstrip('.json')
                setDeckIndex(botHash, customDeckFile, obj_Deck_this)
                setDeckIndex('unity', customDeckFile, obj_Deck_this)
    # 全局 yaml 牌堆
    obj_Deck_this = None
    releaseDir(OlivaDiceCore.data.dataDirRoot + '/unity/extend/deckyaml')
    customDeckDir = OlivaDiceCore.data.dataDirRoot + '/unity/extend/deckyaml'
    fileDeckList = os.listdir(customDeckDir)
    for fileDeckList_this in fileDeckList:
        customDeckFile = fileDeckList_this
        customDeckPath = customDeckDir + '/' + customDeckFile
        obj_Deck_this = None
        try:
            with open(customDeckPath, 'rb') as customDeckPath_f:
                obj_Deck_this = yaml.load(
                    formatUTF8WithBOM(customDeckPath_f.read()).decode('utf-8'),
                    Loader = yaml.FullLoader
                )
        except:
            dictTValue['tInitDataName'] = customDeckFile
            OlivaDiceCore.msgReply.globalLog(
                3,
                OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckDataError'], dictTValue),
                [
                    ('OlivaDice', 'default'),
                    ('Init', 'default')
                ]
            )
        customDeckFile_deckName = customDeckFile
        if customDeckFile_deckName.endswith('.yaml'):
            customDeckFile_deckName = customDeckFile_deckName.rstrip('.yaml')
        if obj_Deck_this != None:
            obj_Deck_this_new = initYamlDeckData(obj_Deck_this, deckName = customDeckFile_deckName)
            if obj_Deck_this_new != None:
                for bot_info_dict_this in OlivaDiceCore.drawCardData.dictDeck:
                    botHash = bot_info_dict_this
                    OlivaDiceCore.drawCardData.dictDeck[botHash].update(obj_Deck_this_new)
                    setDeckIndex(botHash, customDeckFile_deckName, obj_Deck_this_new)
                    setDeckIndex('unity', customDeckFile_deckName, obj_Deck_this_new)
    # 全局 excel 牌堆
    releaseDir(OlivaDiceCore.data.dataDirRoot + '/unity/extend/deckexcel')
    customDeckDir = OlivaDiceCore.data.dataDirRoot + '/unity/extend/deckexcel'
    fileDeckList = os.listdir(customDeckDir)
    for fileDeckList_this in fileDeckList:
        customDeckFile = fileDeckList_this
        customDeckPath = customDeckDir + '/' + customDeckFile
        obj_Deck_this_root = None
        try:
            obj_Deck_this_root = openpyxl.load_workbook(customDeckPath)
        except:
            obj_Deck_this_root = None
            dictTValue['tInitDataName'] = customDeckFile
            OlivaDiceCore.msgReply.globalLog(
                3,
                OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckDataError'], dictTValue),
                [
                    ('OlivaDice', 'default'),
                    ('Init', 'default')
                ]
            )
        if obj_Deck_this_root != None:
            obj_Deck_this_new = None
            try:
                obj_Deck_this_new = initExcelDeckData(obj_Deck_this_root)
            except:
                obj_Deck_this_new = None
                dictTValue['tInitDataName'] = customDeckFile
                OlivaDiceCore.msgReply.globalLog(
                    3,
                    OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckDataError'], dictTValue),
                    [
                        ('OlivaDice', 'default'),
                        ('Init', 'default')
                    ]
                )
            try:
                obj_Deck_this_root.close()
            except:
                pass
            customDeckFile_new = customDeckFile
            if customDeckFile_new.endswith('.xlsx'):
                customDeckFile_new = customDeckFile_new.rstrip('.xlsx')
            elif customDeckFile_new.endswith('.xls'):
                customDeckFile_new = customDeckFile_new.rstrip('.xls')
            if obj_Deck_this_new != None:
                for bot_info_dict_this in OlivaDiceCore.drawCardData.dictDeck:
                    botHash = bot_info_dict_this
                    OlivaDiceCore.drawCardData.dictDeck[botHash].update(obj_Deck_this_new)
                    setDeckIndex(botHash, customDeckFile_new, obj_Deck_this_new)
                    setDeckIndex('unity', customDeckFile_new, obj_Deck_this_new)
    # 全局 牌堆 日志
    if botHash != None:
        obj_Deck_this_count_total = len(OlivaDiceCore.drawCardData.dictDeck[botHash])
        dictTValue['tInitDataCount'] = str(obj_Deck_this_count_total - obj_Deck_this_count_total_init)
        dictTValue['tInitDataCount01'] = str(obj_Deck_this_count_total)
        OlivaDiceCore.msgReply.globalLog(
            2,
            OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckData'], dictTValue),
            [
                ('OlivaDice', 'default'),
                ('Init', 'default')
            ]
        )

    for bot_info_dict_this in OlivaDiceCore.drawCardData.dictDeck:
        # json 牌堆
        botHash = bot_info_dict_this
        releaseDir(OlivaDiceCore.data.dataDirRoot + '/' + botHash)
        releaseDir(OlivaDiceCore.data.dataDirRoot + '/' + botHash + '/extend')
        releaseDir(OlivaDiceCore.data.dataDirRoot + '/' + botHash + '/extend/deckclassic')
        customDeckDir = OlivaDiceCore.data.dataDirRoot + '/' + botHash + '/extend/deckclassic'
        dictTValue['tName'] = botHash
        if botHash in bot_info_dict:
            dictTValue['tName'] = '%s|%s' % (
                bot_info_dict[botHash].platform['platform'],
                str(bot_info_dict[botHash].id)
            )
        fileDeckList = os.listdir(customDeckDir)
        for fileDeckList_this in fileDeckList:
            customDeckFile = fileDeckList_this
            customDeckPath = customDeckDir + '/' + customDeckFile
            obj_Deck_this = None
            try:
                with open(customDeckPath, 'rb') as customDeckPath_f:
                    customDeckPath_fs = formatUTF8WithBOM(customDeckPath_f.read()).decode('utf-8')
                    try:
                        obj_Deck_this = pyjson5.loads(customDeckPath_fs)
                    except:
                        obj_Deck_this = json.loads(customDeckPath_fs)
            except:
                dictTValue['tInitDataName'] = customDeckFile
                OlivaDiceCore.msgReply.globalLog(
                    3,
                    OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckDataError'], dictTValue),
                    [
                        ('OlivaDice', 'default'),
                        ('Init', 'default')
                    ]
                )
            if obj_Deck_this != None:
                OlivaDiceCore.drawCardData.dictDeck[botHash].update(obj_Deck_this)
                if customDeckFile.endswith('.json5'):
                    customDeckFile = customDeckFile.rstrip('.json5')
                elif customDeckFile.endswith('.json'):
                    customDeckFile = customDeckFile.rstrip('.json')
                setDeckIndex(botHash, customDeckFile, obj_Deck_this)
        # yaml 牌堆
        releaseDir(OlivaDiceCore.data.dataDirRoot + '/' + botHash + '/extend/deckyaml')
        customDeckDir = OlivaDiceCore.data.dataDirRoot + '/' + botHash + '/extend/deckyaml'
        fileDeckList = os.listdir(customDeckDir)
        for fileDeckList_this in fileDeckList:
            customDeckFile = fileDeckList_this
            customDeckPath = customDeckDir + '/' + customDeckFile
            obj_Deck_this = None
            try:
                with open(customDeckPath, 'rb') as customDeckPath_f:
                    obj_Deck_this = yaml.load(
                        formatUTF8WithBOM(customDeckPath_f.read()).decode('utf-8'),
                        Loader = yaml.FullLoader
                    )
            except:
                dictTValue['tInitDataName'] = customDeckFile
                OlivaDiceCore.msgReply.globalLog(
                    3,
                    OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckDataError'], dictTValue),
                    [
                        ('OlivaDice', 'default'),
                        ('Init', 'default')
                    ]
                )
            customDeckFile_deckName = customDeckFile
            if customDeckFile_deckName.endswith('.yaml'):
                customDeckFile_deckName = customDeckFile_deckName.rstrip('.yaml')
            if obj_Deck_this != None:
                obj_Deck_this_new = initYamlDeckData(obj_Deck_this, deckName = customDeckFile_deckName)
                OlivaDiceCore.drawCardData.dictDeck[botHash].update(obj_Deck_this_new)
                setDeckIndex(botHash, customDeckFile_deckName, obj_Deck_this_new)
        # 全局 excel 牌堆
        releaseDir(OlivaDiceCore.data.dataDirRoot + '/' + botHash + '/extend/deckexcel')
        customDeckDir = OlivaDiceCore.data.dataDirRoot + '/' + botHash + '/extend/deckexcel'
        fileDeckList = os.listdir(customDeckDir)
        for fileDeckList_this in fileDeckList:
            customDeckFile = fileDeckList_this
            customDeckPath = customDeckDir + '/' + customDeckFile
            obj_Deck_this_root = None
            try:
                obj_Deck_this_root = openpyxl.load_workbook(customDeckPath)
            except:
                obj_Deck_this_root = None
                dictTValue['tInitDataName'] = customDeckFile
                OlivaDiceCore.msgReply.globalLog(
                    3,
                    OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckDataError'], dictTValue),
                    [
                        ('OlivaDice', 'default'),
                        ('Init', 'default')
                    ]
                )
            if obj_Deck_this_root != None:
                obj_Deck_this_new = None
                try:
                    obj_Deck_this_new = initExcelDeckData(obj_Deck_this_root)
                except:
                    obj_Deck_this_new = None
                    dictTValue['tInitDataName'] = customDeckFile
                    OlivaDiceCore.msgReply.globalLog(
                        3,
                        OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckDataError'], dictTValue),
                        [
                            ('OlivaDice', 'default'),
                            ('Init', 'default')
                        ]
                    )
                try:
                    obj_Deck_this_root.close()
                except:
                    pass
                customDeckFile_new = customDeckFile
                if customDeckFile_new.endswith('.xlsx'):
                    customDeckFile_new = customDeckFile_new.rstrip('.xlsx')
                elif customDeckFile_new.endswith('.xls'):
                    customDeckFile_new = customDeckFile_new.rstrip('.xls')
                if obj_Deck_this_new != None:
                    OlivaDiceCore.drawCardData.dictDeck[botHash].update(obj_Deck_this_new)
                    setDeckIndex(botHash, customDeckFile_new, obj_Deck_this_new)
        # 日志
        obj_Deck_this_count = len(OlivaDiceCore.drawCardData.dictDeck[botHash])
        dictTValue['tInitDataCount'] = str(obj_Deck_this_count - obj_Deck_this_count_total)
        dictTValue['tInitDataCount01'] = str(obj_Deck_this_count)
        OlivaDiceCore.msgReply.globalLog(
            2,
            OlivaDiceCore.msgCustomManager.formatReplySTRConst(dictStrConst['strInitDeckData'], dictTValue),
            [
                ('OlivaDice', 'default'),
                ('Init', 'default')
            ]
        )
    initDeckHelp(bot_info_dict = bot_info_dict)

def initExcelDeckData(data):
    res = {}
    for obj_Deck_this_name in data.sheetnames:
        obj_Deck_this = None
        obj_Deck_this = data.get_sheet_by_name(obj_Deck_this_name)
        checkIndex = []
        deckData = []
        flag_first = True
        for data_row in obj_Deck_this:
            data_row_offset = 0
            checkData = {
                'Content': None,
                'Weight': None,
                'Redraw': None,
                'Finalize': None
            }
            for data_cell in data_row:
                data_row_value = data_cell.value
                if flag_first:
                    checkIndex.append(data_row_value)
                else:
                    if len(checkIndex) > data_row_offset:
                        dataType = checkIndex[data_row_offset]
                        if dataType in checkData and data_row_value not in [None, '']:
                            checkData[dataType] = data_row_value
                data_row_offset += 1
            if flag_first:
                flag_first = False
            else:
                if checkData['Content'] != None:
                    tmp_deckData_this = checkData['Content']
                    tmp_deckData_this = re.sub(r'DRAW\(\s*(.+?)\s*,\s*0\s*\)', r'{\1}', tmp_deckData_this)
                    tmp_deckData_this = re.sub(r'DRAW\(\s*(.+?)\s*,\s*1\s*\)', r'{%\1}', tmp_deckData_this)
                    tmp_deckData_this = re.sub(r'DRAW\(\s*(.+?)\s*\)', r'{\1}', tmp_deckData_this)
                    if checkData['Weight'] != None:
                        tmp_deckData_this = '::%s::%s' % (
                            str(checkData['Weight']),
                            tmp_deckData_this
                        )
                    deckData.append(tmp_deckData_this)
        if len(deckData) > 0:
            res[obj_Deck_this_name] = deckData
    return res

def initYamlDeckData(data:dict, deckName = None):
    name = deckName
    includes = ['default']
    default = None
    keyList = ['name', 'author', 'version', 'command', 'desc', 'includes']
    res = {}
    if name == None and 'name' in data:
        name = data['name']
    if 'includes' in data:
        includes = data['includes']
    if 'default' in data:
        default = data['default']
    if 'default' not in includes:
        includes.append('default')
    if None not in [name, includes, default]:
        for key in data:
            keyNew = '%s:%s' % (name, key)
            dataNew = []
            if key not in keyList:
                for dataThis in data[key]:
                    dataThisNew = str(dataThis)
                    dataThisNew = re.sub(r'\{%(.+?)\}', r'{_' + name + r':\1}', dataThisNew)
                    dataThisNew = re.sub(r'\{\$(.+?)\}', r'{%_' + name + r':\1}', dataThisNew)
                    for keyIncThis in includes:
                        keyIncThisNew = '%s:%s' % (name, keyIncThis)
                        dataThisNew = dataThisNew.replace('{_%s}' % keyIncThisNew, '{%s}' % keyIncThisNew)
                        dataThisNew = dataThisNew.replace('{%%_%s}' % keyIncThisNew, '{%%%s}' % keyIncThisNew)
                    dataNew.append(dataThisNew)
                if key in includes:
                    res[keyNew] = dataNew
                else:
                    res['_%s' % keyNew] = dataNew
                if 'default' == key:
                    res[name] = dataNew
    return res

def getDrawDeck(key_str, bot_hash, count = 1, valDict = None):
    dictTValue = OlivaDiceCore.msgCustom.dictTValue.copy()
    dictStrCustom = OlivaDiceCore.msgCustom.dictStrCustom
    tmp_reply_str = None
    
    # 应用重定向逻辑（仅用于牌堆数据，自定义回复不重定向）
    redirected_bot_hash = OlivaDiceCore.userConfig.getRedirectedBotHash(bot_hash)
    
    if bot_hash in OlivaDiceCore.msgCustom.dictStrCustomDict:
        dictStrCustom = OlivaDiceCore.msgCustom.dictStrCustomDict[bot_hash]
    if redirected_bot_hash in OlivaDiceCore.drawCardData.dictDeck:
        if key_str in OlivaDiceCore.drawCardData.dictDeck[redirected_bot_hash]:
            if count >= 1 and count <= 10:
                tmp_for_list = range(count)
                tmp_card_list = []
                plugin_event = None
                dictTValue = None
                if valDict != None and 'dictTValue' in valDict and 'vValDict' in valDict['dictTValue'] and 'vPluginEvent' in valDict['dictTValue']['vValDict']:
                    plugin_event = valDict['dictTValue']['vValDict']['vPluginEvent']
                if valDict != None and 'dictTValue' in valDict:
                    dictTValue = valDict['dictTValue']
                for tmp_for_list_this in tmp_for_list:
                    tmp_draw_str = draw(key_str, bot_hash, mark_dict = None, plugin_event = plugin_event)
                    if tmp_draw_str != None and type(tmp_draw_str) == str:
                        tmp_card_list.append(tmp_draw_str)
                dictTValue['tDrawDeckResult'] = '\n'.join(tmp_card_list)
                tmp_reply_str = OlivaDiceCore.msgCustomManager.formatReplySTR(dictStrCustom['strDrawDeck'], dictTValue)
                return tmp_reply_str
    tmp_recommend_list = []
    if OlivaDiceCore.console.getConsoleSwitchByHash(
        'drawRecommendMode',
        bot_hash
    ) == 1:
        tmp_recommend_list = getDeckRecommend(key_str, bot_hash)
    if type(tmp_recommend_list) == list:
        if len(tmp_recommend_list) > 0:
            tmp_recommend_str = '\n'.join(
                ['[.draw %s]' % tmp_recommend_list_this for tmp_recommend_list_this in tmp_recommend_list]
            )
            dictTValue['tDrawDeckResult'] = tmp_recommend_str
            tmp_reply_str = OlivaDiceCore.msgCustomManager.formatReplySTR(dictStrCustom['strDrawDeckRecommend'], dictTValue)
        else:
            tmp_reply_str = OlivaDiceCore.msgCustomManager.formatReplySTR(dictStrCustom['strDrawDeckNotFound'], dictTValue)
    return tmp_reply_str

def getDeckRecommend(key_str:str, bot_hash:str):
    res = []
    tmp_RecommendRank_list = []
    
    # 应用重定向逻辑
    redirected_bot_hash = OlivaDiceCore.userConfig.getRedirectedBotHash(bot_hash)
    
    if redirected_bot_hash in OlivaDiceCore.drawCardData.dictDeck:
        for dictDeck_this in OlivaDiceCore.drawCardData.dictDeck[redirected_bot_hash]:
            tmp_RecommendRank_list.append([
                OlivaDiceCore.helpDoc.getRecommendRank(
                    key_str,
                    dictDeck_this
                ),
                dictDeck_this
            ])
        tmp_RecommendRank_list.sort(key = lambda x : x[0])
    tmp_count_max = min(8, len(tmp_RecommendRank_list))
    count = 0
    while count < tmp_count_max:
        if tmp_RecommendRank_list[count][0] < 1000:
            if len(tmp_RecommendRank_list[count][1]) < 25:
                if len(tmp_RecommendRank_list[count][1]) > 0 and tmp_RecommendRank_list[count][1][0] != '_':
                    res.append(tmp_RecommendRank_list[count][1])
        count += 1
    return res

def draw(key_str:str, bot_hash:str, flag_need_give_back:bool = True, mark_dict:'dict|None' = None, plugin_event:'OlivOS.API.Event|None' = None):
    tmp_reply_str = None
    tmp_deck_this = []
    tmp_deck_this_len = 1
    tmp_deck_this_hit = 0
    
    # 应用重定向逻辑
    redirected_bot_hash = OlivaDiceCore.userConfig.getRedirectedBotHash(bot_hash)
    
    if redirected_bot_hash in OlivaDiceCore.drawCardData.dictDeck:
        if key_str in OlivaDiceCore.drawCardData.dictDeck[redirected_bot_hash]:
            if mark_dict == None:
                tmp_mark_dict = {}
            else:
                tmp_mark_dict = mark_dict
            if key_str in tmp_mark_dict:
                if len(tmp_mark_dict[key_str]) <= 0:
                    tmp_mark_dict[key_str] = initCloneDeckList(OlivaDiceCore.drawCardData.dictDeck[redirected_bot_hash][key_str])
            else:
                tmp_mark_dict[key_str] = initCloneDeckList(OlivaDiceCore.drawCardData.dictDeck[redirected_bot_hash][key_str])
            tmp_deck_this = tmp_mark_dict[key_str]
            if type(tmp_deck_this) == list:
                tmp_deck_this_len = len(tmp_deck_this)
                if tmp_deck_this_len > 0:
                    tmp_deck_this_hit = random.randint(0, tmp_deck_this_len - 1)
                    tmp_reply_str = tmp_deck_this[tmp_deck_this_hit]
                    if not flag_need_give_back:
                        tmp_deck_this.pop(tmp_deck_this_hit)
                    flag_need_roll = True
                    tmp_mark_left = -1
                    tmp_mark_right = -1
                    tmp_base_offset = 0
                    tmp_reply_str_2 = tmp_reply_str
                    tmp_reply_str = ''
                    while flag_need_roll:
                        tmp_reply_str_1 = None
                        tmp_mark_left = tmp_reply_str_2.find('{', tmp_base_offset)
                        flag_need_give_back_2 = False
                        if tmp_mark_left != -1:
                            tmp_mark_right = tmp_reply_str_2.find('}', tmp_mark_left)
                        if tmp_mark_left != -1 and tmp_mark_right != -1:
                            tmp_reply_str += tmp_reply_str_2[tmp_base_offset:tmp_mark_left]
                            if tmp_mark_right - tmp_mark_left > 2:
                                if tmp_reply_str_2[tmp_mark_left + 1] == '%':
                                    tmp_reply_str_1 = tmp_reply_str_2[tmp_mark_left + 2:tmp_mark_right]
                                    flag_need_give_back_2 = True
                                else:
                                    tmp_reply_str_1 = tmp_reply_str_2[tmp_mark_left + 1:tmp_mark_right]
                            else:
                                tmp_reply_str_1 = tmp_reply_str_2[tmp_mark_left + 1:tmp_mark_right]
                            tmp_reply_str_1 = draw(tmp_reply_str_1, bot_hash, flag_need_give_back_2, tmp_mark_dict, plugin_event = plugin_event)
                            if tmp_reply_str_1 != None:
                                tmp_reply_str += tmp_reply_str_1
                            else:
                                tmp_reply_str += tmp_reply_str_2[tmp_mark_left:tmp_mark_right + 1]
                            tmp_base_offset = tmp_mark_right + 1
                        else:
                            tmp_reply_str += tmp_reply_str_2[tmp_base_offset:]
                            flag_need_roll = False
                        tmp_mark_left = -1
                        tmp_mark_right = -1
                    flag_need_roll = True
                    tmp_mark_left = -1
                    tmp_mark_right = -1
                    tmp_base_offset = 0
                    tmp_reply_str_2 = tmp_reply_str
                    tmp_reply_str = ''
                    while flag_need_roll:
                        tmp_reply_str_1 = None
                        tmp_mark_left = tmp_reply_str_2.find('[', tmp_base_offset)
                        flag_need_give_back_2 = False
                        if tmp_mark_left != -1:
                            tmp_mark_right = tmp_reply_str_2.find(']', tmp_mark_left)
                        if tmp_mark_left != -1 and tmp_mark_right != -1:
                            tmp_reply_str += tmp_reply_str_2[tmp_base_offset:tmp_mark_left]
                            tmp_reply_str_1 = tmp_reply_str_2[tmp_mark_left + 1:tmp_mark_right]
                            tmp_rd = OlivaDiceCore.onedice.RD(tmp_reply_str_1)
                            tmp_rd.roll()
                            if tmp_rd.resError == None:
                                tmp_reply_str += str(tmp_rd.resInt)
                            else:
                                tmp_reply_str += tmp_reply_str_2[tmp_mark_left:tmp_mark_right + 1]
                            tmp_base_offset = tmp_mark_right + 1
                        else:
                            tmp_reply_str += tmp_reply_str_2[tmp_base_offset:]
                            flag_need_roll = False
                        tmp_mark_left = -1
                        tmp_mark_right = -1
    dictTValue = {}
    if tmp_reply_str != None and plugin_event != None:
        tmp_reply_str = OlivaDiceCore.crossHook.dictHookFunc['drawFormatHook'](tmp_reply_str, plugin_event)
        dictTValue = dictTValueGet(plugin_event)
    if tmp_reply_str != None and type(tmp_reply_str) == str:
        tmp_reply_str = reMappingDrawFormat(tmp_reply_str)
        if dictTValue != None:
            tmp_reply_str = OlivaDiceCore.msgCustomManager.formatReplySTR(
                tmp_reply_str, dictTValue, flagCross = False, flagSplit = False
            )
    return tmp_reply_str

def dictTValueGet(plugin_event):
    res = {}
    if 'name' in plugin_event.data.sender:
        res['tName'] = plugin_event.data.sender['name']
    return res

def initCloneDeckList(src_deck):
    res_deck = []
    for src_deck_this in src_deck:
        tmp_mark_left = -1
        tmp_mark_right = -1
        tmp_reply_str_1 = None
        tmp_reply_str_2 = src_deck_this
        if len(tmp_reply_str_2) > 4:
            if tmp_reply_str_2[:2] == '::':
                tmp_mark_left = 0
            if tmp_mark_left != -1:
                tmp_mark_right = tmp_reply_str_2.find('::', 2)
        if tmp_mark_left != -1 and tmp_mark_right != -1:
            tmp_reply_str_1 = tmp_reply_str_2[tmp_mark_left + 2:tmp_mark_right]
            if len(tmp_reply_str_1) > 3 and tmp_reply_str_1[0] == '[' and tmp_reply_str_1[-1] == ']':
                tmp_reply_str_1 = tmp_reply_str_1[1:-1]
                tmp_rd = OlivaDiceCore.onedice.RD(tmp_reply_str_1)
                tmp_rd.roll()
                if tmp_rd.resError == None:
                    if tmp_rd.resInt > 0:
                        if tmp_mark_right + 2 < len(tmp_reply_str_2):
                            res_deck += [tmp_reply_str_2[tmp_mark_right + 2:]] * tmp_rd.resInt
                else:
                    res_deck += [src_deck_this]
            elif tmp_reply_str_1.isdigit():
                if tmp_mark_right + 2 < len(tmp_reply_str_2):
                    res_deck += [tmp_reply_str_2[tmp_mark_right + 2:]] * int(tmp_reply_str_1)
                else:
                    res_deck += [src_deck_this]
            else:
                res_deck += [src_deck_this]
        else:
            res_deck += [src_deck_this]
    return res_deck

def releaseDir(dir_path):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
